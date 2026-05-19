#!/usr/bin/env python3
"""
Raw PDF extraction review workbook.

Extracts a PDF using Docling and writes a human-reviewable Excel workbook.
No LLM calls. No schema mapping. No CSV staging for import. Local only.

Usage:
    python scripts/review/raw_pdf_review_workbook.py \
        --input "C:\\path\\to\\document.pdf" \
        [--pages "5-7"] \
        [--run-name "my-run"] \
        [--no-image-processing] \
        [--allow-partial]

Output folder: .local/extraction-runs/<run-name>/

Activate .venv-docling before running.
"""

import argparse
import csv
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SAFE_OUTPUT_ROOT = Path(".local/extraction-runs")

HEADING_LABELS = {"title", "section_header"}

# DocItemLabel values that carry prose content (non-table, non-heading)
CHUNK_LABELS = {
    "paragraph", "caption", "footnote", "list_item",
    "code", "formula", "text", "picture", "page_header", "page_footer",
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
REVIEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
REVIEW_FONT = Font(color="7F6000", italic=True, size=10)

TEXT_PREVIEW_CHARS = 600
NEARBY_PREVIEW_CHARS = 300
MAX_COL_WIDTH = 80
MIN_COL_WIDTH = 10


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def make_safe_name(text: str) -> str:
    return re.sub(r"[^\w\-]", "_", text)[:60]


def resolve_output_dir(run_name, stem: str, ts: str) -> Path:
    name = run_name or f"{make_safe_name(stem)}_{ts}"
    out = SAFE_OUTPUT_ROOT / name
    resolved_root = SAFE_OUTPUT_ROOT.resolve()
    resolved_out = out.resolve()
    try:
        resolved_out.relative_to(resolved_root)
    except ValueError:
        sys.exit(
            f"[ERROR] --run-name resolves outside {SAFE_OUTPUT_ROOT}/.\nGot: {out}"
        )
    return out


def label_str(item) -> str:
    label = getattr(item, "label", None)
    if label is None:
        return "unknown"
    if hasattr(label, "value"):
        return str(label.value)
    return str(label)


def is_heading(item) -> bool:
    return label_str(item).lower() in HEADING_LABELS


def get_page_no(item) -> int | None:
    prov = getattr(item, "prov", None) or []
    if prov:
        return getattr(prov[0], "page_no", None)
    return None


def get_bbox_sort_key(item) -> float:
    """Return a vertical sort key for reading-order within a page.

    Docling uses PDF coordinates (origin bottom-left), so higher t = higher
    on page. We negate so that items near the top of the page sort first.
    """
    prov = getattr(item, "prov", None) or []
    if prov:
        bbox = getattr(prov[0], "bbox", None)
        if bbox is not None:
            t = getattr(bbox, "t", None)
            if t is not None:
                return -float(t)
    return 0.0


def safe_text(item) -> str:
    return (getattr(item, "text", "") or "").strip()


def doc_to_json_str(doc) -> str:
    try:
        return doc.model_dump_json(indent=2)
    except Exception:
        try:
            return json.dumps(doc.model_dump(), indent=2, default=str)
        except Exception as exc:
            return json.dumps({"error": str(exc)}, indent=2)


def truncate(s: str, n: int) -> str:
    s = s.strip()
    if len(s) <= n:
        return s
    return s[:n] + "…"


# ---------------------------------------------------------------------------
# Page selection & image rendering
# ---------------------------------------------------------------------------

def parse_pages_arg(pages_str: str) -> list[int]:
    """Parse a pages argument like '5-7' or '5,6,7' into a sorted list of 1-indexed page numbers."""
    pages = set()
    for part in pages_str.split(","):
        part = part.strip()
        if "-" in part:
            a, b = part.split("-", 1)
            pages.update(range(int(a), int(b) + 1))
        else:
            pages.add(int(part))
    return sorted(pages)


def extract_pages_to_pdf(input_path: Path, page_nos: list[int], out_path: Path) -> Path:
    """Extract specific pages (1-indexed) from a PDF into a new PDF file using pypdfium2."""
    import pypdfium2  # noqa: PLC0415
    src = pypdfium2.PdfDocument(str(input_path))
    dst = pypdfium2.PdfDocument.new()
    zero_indexed = [p - 1 for p in page_nos]
    dst.import_pages(src, pages=zero_indexed)
    dst.save(str(out_path))
    src.close()
    dst.close()
    print(f"[workbook] Extracted pages {page_nos} -> {out_path.name}")
    return out_path


def render_page_images(input_path: Path, page_nos: list[int], images_dir: Path, scale: float = 2.0) -> list[Path]:
    """Render PDF pages (1-indexed) as PNG files. Returns list of written PNG paths."""
    import pypdfium2  # noqa: PLC0415
    images_dir.mkdir(parents=True, exist_ok=True)
    pdf = pypdfium2.PdfDocument(str(input_path))
    written = []
    for pn in page_nos:
        page = pdf[pn - 1]
        bitmap = page.render(scale=scale)
        pil_img = bitmap.to_pil()
        png_path = images_dir / f"page_{pn:02d}.png"
        pil_img.save(str(png_path))
        written.append(png_path)
        print(f"[workbook] Rendered page {pn} -> {png_path.name} ({pil_img.size[0]}x{pil_img.size[1]}px)")
    pdf.close()
    return written


# ---------------------------------------------------------------------------
# Docling extraction
# ---------------------------------------------------------------------------

def run_docling(input_path: Path, no_image_processing: bool, allow_partial: bool, out_dir: Path):
    """Run Docling and return the document object."""
    from docling.datamodel.base_models import InputFormat  # noqa: PLC0415
    from docling.document_converter import DocumentConverter, PdfFormatOption  # noqa: PLC0415
    from docling.datamodel.pipeline_options import PdfPipelineOptions  # noqa: PLC0415

    pipeline_options = PdfPipelineOptions()
    pipeline_options.do_ocr = False

    if no_image_processing:
        from docling.backend.docling_parse_backend import DoclingParseDocumentBackend  # noqa: PLC0415
        pipeline_options.do_table_structure = False
        converter = DocumentConverter(
            format_options={
                InputFormat.PDF: PdfFormatOption(
                    pipeline_options=pipeline_options,
                    backend=DoclingParseDocumentBackend,
                )
            }
        )
        print("[docling] Using DoclingParse backend (no image rendering, no table vision).")
    else:
        converter = DocumentConverter(
            format_options={
                InputFormat.PDF: PdfFormatOption(pipeline_options=pipeline_options)
            }
        )
        print("[docling] Using default backend (table structure detection enabled).")
        print("[docling] Note: first run may download AI models (~1-2 GB).")

    print(f"[docling] Converting: {input_path.name} ...")
    result = converter.convert(str(input_path))
    doc = result.document

    # Coverage check
    page_count = len(getattr(doc, "pages", None) or {})
    extracted = set()
    for item in list(getattr(doc, "texts", []) or []) + list(getattr(doc, "tables", []) or []):
        for prov in getattr(item, "prov", []) or []:
            pn = getattr(prov, "page_no", None)
            if pn is not None:
                extracted.add(int(pn))
    extracted_max = max(extracted) if extracted else None

    coverage_ok = True
    if not allow_partial and page_count and extracted_max is not None:
        coverage_pct = extracted_max / page_count
        if coverage_pct < 0.9:
            coverage_ok = False
            print(
                f"\n[WARNING] Docling stopped at page {extracted_max} of {page_count} "
                f"({coverage_pct:.0%} coverage). Pages {extracted_max + 1}–{page_count} may be missing.\n"
                f"Run with --allow-partial to suppress this warning.\n"
            )

    return doc, page_count, len(extracted), extracted_max, coverage_ok


# ---------------------------------------------------------------------------
# Table structure helpers
# ---------------------------------------------------------------------------

def _grid_has_content(tbl) -> bool:
    """Return True if the table grid contains at least one non-empty cell."""
    try:
        for row in tbl.data.grid:
            for cell in row:
                if (cell.text or "").strip():
                    return True
    except Exception:
        pass
    return False


def _resolve_table_group_children(doc, tbl) -> list:
    """For a table with an empty grid, return the text items from the associated group.

    When Docling is run without table structure detection it finds the table
    *region* but doesn't parse rows/columns. Instead the cells are stored as
    plain text items inside a GroupItem whose self_ref is pointed to by
    table_cells[0].ref.cref.
    """
    try:
        cell0 = tbl.data.table_cells[0]
        cref = cell0.ref.cref  # e.g. '#/groups/3'
        if "groups" not in cref:
            return []
        g_idx = int(cref.split("/")[-1])
        groups = list(getattr(doc, "groups", []) or [])
        if g_idx >= len(groups):
            return []
        group = groups[g_idx]
        texts = list(getattr(doc, "texts", []) or [])
        result = []
        for child_ref in getattr(group, "children", []) or []:
            child_cref = getattr(child_ref, "cref", "")
            if "texts" in child_cref:
                t_idx = int(child_cref.split("/")[-1])
                if t_idx < len(texts):
                    result.append(texts[t_idx])
        return result
    except Exception:
        return []


def build_table_child_refs(doc) -> set:
    """Return a set of self_ref strings for text items that live inside a table group.

    These items are already represented in the table sheets — excluding them
    from sheet 4 (raw_chunks) avoids double-counting.
    """
    child_refs: set[str] = set()
    tables = list(getattr(doc, "tables", []) or [])
    for tbl in tables:
        if _grid_has_content(tbl):
            continue
        for item in _resolve_table_group_children(doc, tbl):
            ref = getattr(item, "self_ref", None)
            if ref:
                child_refs.add(ref)
    return child_refs


# ---------------------------------------------------------------------------
# Data builders
# ---------------------------------------------------------------------------

def build_page_items(doc) -> dict:
    """Return dict: page_no (int) -> list of (sort_key, item) in reading order."""
    page_items: dict[int, list] = {}

    for item in list(getattr(doc, "texts", []) or []):
        pn = get_page_no(item)
        if pn is None:
            continue
        page_items.setdefault(pn, []).append((get_bbox_sort_key(item), item))

    for item in list(getattr(doc, "tables", []) or []):
        pn = get_page_no(item)
        if pn is None:
            continue
        page_items.setdefault(pn, []).append((get_bbox_sort_key(item), item))

    # Sort each page by reading order
    for pn in page_items:
        page_items[pn].sort(key=lambda x: x[0])

    return page_items


def build_page_heading_map(page_items: dict) -> dict:
    """Return dict: page_no -> list of heading items in reading order."""
    result = {}
    for pn, items in page_items.items():
        headings = [item for _, item in items if is_heading(item)]
        result[pn] = headings
    return result


def sheet_1_source_pages(doc, page_items: dict, page_heading_map: dict) -> pd.DataFrame:
    rows = []
    all_pages = set(page_items.keys())

    # Also include pages declared by Docling even if no content was extracted
    doc_pages = getattr(doc, "pages", None) or {}
    try:
        for pn in doc_pages:
            all_pages.add(int(pn))
    except Exception:
        pass

    for pn in sorted(all_pages):
        items_on_page = [item for _, item in page_items.get(pn, [])]
        headings = page_heading_map.get(pn, [])

        h1 = safe_text(headings[0]) if len(headings) > 0 else ""
        h2 = safe_text(headings[1]) if len(headings) > 1 else ""

        # Text preview: first non-heading text
        preview_parts = []
        for item in items_on_page:
            if not is_heading(item) and hasattr(item, "text"):
                t = safe_text(item)
                if t:
                    preview_parts.append(t)
            if sum(len(p) for p in preview_parts) >= TEXT_PREVIEW_CHARS:
                break
        preview = truncate(" ".join(preview_parts), TEXT_PREVIEW_CHARS)

        rows.append({
            "page_number": pn,
            "page_label": f"Page {pn}",
            "detected_heading": h1,
            "detected_subheading": h2,
            "text_preview": preview,
            "human_notes": "",
            "use_for_extraction": "",
        })

    return pd.DataFrame(rows, columns=[
        "page_number", "page_label", "detected_heading", "detected_subheading",
        "text_preview", "human_notes", "use_for_extraction",
    ])


def sheet_2_raw_tables(doc) -> pd.DataFrame:
    rows = []
    tables = list(getattr(doc, "tables", []) or [])

    for tbl_idx, tbl in enumerate(tables, start=1):
        pn = get_page_no(tbl)

        if _grid_has_content(tbl):
            # Normal path: table structure detection produced a real grid
            try:
                for row_idx, row in enumerate(tbl.data.grid):
                    for col_idx, cell in enumerate(row):
                        cell_text = (cell.text or "").strip()
                        rows.append({
                            "page_number": pn,
                            "table_number": tbl_idx,
                            "row_number": row_idx,
                            "column_number": col_idx,
                            "cell_text": cell_text,
                            "human_notes": "",
                            "use_for_extraction": "",
                        })
            except Exception:
                pass
        else:
            # Fallback: no table structure detected — list group children as flat cells
            children = _resolve_table_group_children(doc, tbl)
            for child_idx, item in enumerate(children):
                cell_text = safe_text(item)
                rows.append({
                    "page_number": pn,
                    "table_number": tbl_idx,
                    "row_number": child_idx,
                    "column_number": 0,
                    "cell_text": cell_text,
                    "human_notes": "[no grid — table structure detection was off]",
                    "use_for_extraction": "",
                })

    return pd.DataFrame(rows, columns=[
        "page_number", "table_number", "row_number", "column_number",
        "cell_text", "human_notes", "use_for_extraction",
    ])


def sheet_3_raw_table_rows(doc) -> pd.DataFrame:
    rows = []
    tables = list(getattr(doc, "tables", []) or [])

    def _make_row(pn, tbl_idx, row_idx, cells, note=""):
        raw_row_text = " | ".join(cells)
        row_dict: dict = {
            "page_number": pn,
            "table_number": tbl_idx,
            "row_number": row_idx,
            "raw_row_text": raw_row_text,
        }
        for i in range(12):
            row_dict[f"cell_{i + 1}"] = cells[i] if i < len(cells) else ""
        row_dict["human_notes"] = note
        row_dict["use_for_extraction"] = ""
        return row_dict

    for tbl_idx, tbl in enumerate(tables, start=1):
        pn = get_page_no(tbl)

        if _grid_has_content(tbl):
            # Normal path: table structure detection produced a real grid
            try:
                for row_idx, row in enumerate(tbl.data.grid):
                    cells = [(cell.text or "").strip() for cell in row]
                    rows.append(_make_row(pn, tbl_idx, row_idx, cells))
            except Exception:
                pass
        else:
            # Fallback: show group children as individual single-cell rows
            children = _resolve_table_group_children(doc, tbl)
            note = "[no grid — table structure detection was off; each row is one text fragment]"
            for child_idx, item in enumerate(children):
                rows.append(_make_row(pn, tbl_idx, child_idx, [safe_text(item)], note))

    cols = (
        ["page_number", "table_number", "row_number", "raw_row_text"]
        + [f"cell_{i}" for i in range(1, 13)]
        + ["human_notes", "use_for_extraction"]
    )
    return pd.DataFrame(rows, columns=cols)


def sheet_4_raw_chunks(doc, page_items: dict, page_heading_map: dict, table_child_refs: set) -> pd.DataFrame:
    rows = []
    chunk_id = 0

    for pn in sorted(page_items.keys()):
        items = [item for _, item in page_items[pn]]
        current_heading = ""
        text_order = 0

        for item in items:
            lbl = label_str(item)

            if is_heading(item):
                current_heading = safe_text(item)
                continue  # headings go into sheet 5, not sheet 4

            # Skip table items (TableItem) — they have their own sheets
            if not hasattr(item, "text"):
                continue

            # Skip text items that are children of a table group — already in sheets 2/3
            item_ref = getattr(item, "self_ref", None)
            if item_ref and item_ref in table_child_refs:
                continue

            txt = safe_text(item)
            if not txt:
                continue

            chunk_id += 1
            rows.append({
                "chunk_id": chunk_id,
                "page_number": pn,
                "detected_heading": current_heading,
                "chunk_type": lbl,
                "text_order_on_page": text_order,
                "chunk_text": txt,
                "human_notes": "",
                "use_for_extraction": "",
            })
            text_order += 1

    return pd.DataFrame(rows, columns=[
        "chunk_id", "page_number", "detected_heading", "chunk_type",
        "text_order_on_page", "chunk_text", "human_notes", "use_for_extraction",
    ])


def sheet_5_headings_index(doc, page_items: dict) -> pd.DataFrame:
    rows = []

    for pn in sorted(page_items.keys()):
        items = [item for _, item in page_items[pn]]

        for idx, item in enumerate(items):
            if not is_heading(item):
                continue

            lbl = label_str(item).lower()
            if lbl == "title":
                level = 1
            elif lbl == "section_header":
                level = 2
            else:
                level = 3

            # Nearby text: next non-heading item on the same page
            nearby = ""
            for nxt in items[idx + 1:]:
                if is_heading(nxt):
                    break
                if hasattr(nxt, "text"):
                    t = safe_text(nxt)
                    if t:
                        nearby = truncate(t, NEARBY_PREVIEW_CHARS)
                        break

            rows.append({
                "page_number": pn,
                "heading_level": level,
                "heading_text": safe_text(item),
                "nearby_text_preview": nearby,
                "human_notes": "",
                "use_for_extraction": "",
            })

    return pd.DataFrame(rows, columns=[
        "page_number", "heading_level", "heading_text",
        "nearby_text_preview", "human_notes", "use_for_extraction",
    ])


def sheet_3b_bbox_rows(doc, page_offset: int = 0) -> pd.DataFrame:
    """Reconstruct rows from raw text fragments using bbox y-coordinate banding.

    When Docling detects a table region as a picture (complex PDF layouts), text
    items are still extracted with accurate bbox coordinates. This sheet groups
    those fragments by their vertical position to approximate the original row
    structure, without making any schema assumptions.

    Row grouping tolerance: items within Y_BAND_TOLERANCE units of each other
    (in PDF point coordinates) are treated as the same row.
    """
    Y_BAND_TOLERANCE = 6.0  # PDF points — adjust if rows merge or split unexpectedly
    MAX_COLS = 12

    texts = list(getattr(doc, "texts", []) or [])
    rows_out = []

    # Collect text items per page with their bbox
    page_map: dict[int, list] = {}
    for item in texts:
        prov = getattr(item, "prov", None) or []
        if not prov:
            continue
        pn = getattr(prov[0], "page_no", None)
        if pn is None:
            continue
        bbox = getattr(prov[0], "bbox", None)
        if bbox is None:
            continue
        t = getattr(bbox, "t", None)
        l = getattr(bbox, "l", None)
        if t is None or l is None:
            continue
        txt = (getattr(item, "text", "") or "").strip()
        if not txt:
            continue
        page_map.setdefault(pn, []).append((float(t), float(l), txt))

    for pn in sorted(page_map.keys()):
        items = sorted(page_map[pn], key=lambda x: (-x[0], x[1]))  # top→bottom, left→right

        # Band items into rows by y-coordinate proximity
        bands: list[list] = []
        for t_val, l_val, txt in items:
            placed = False
            for band in bands:
                band_t = band[0][0]  # representative t of this band
                if abs(t_val - band_t) <= Y_BAND_TOLERANCE:
                    band.append((t_val, l_val, txt))
                    placed = True
                    break
            if not placed:
                bands.append([(t_val, l_val, txt)])

        # Sort each band left-to-right, then output as a row
        for row_idx, band in enumerate(bands):
            band.sort(key=lambda x: x[1])  # sort by l (left-to-right)
            cells = [txt for _, _, txt in band]
            raw_text = " | ".join(cells)
            row_dict: dict = {
                "page_number": pn + page_offset,
                "row_number": row_idx,
                "raw_text": raw_text,
            }
            for i in range(MAX_COLS):
                row_dict[f"col_{i + 1}"] = cells[i] if i < len(cells) else ""
            row_dict["human_notes"] = ""
            row_dict["use_for_extraction"] = ""
            rows_out.append(row_dict)

    cols = (
        ["page_number", "row_number", "raw_text"]
        + [f"col_{i}" for i in range(1, MAX_COLS + 1)]
        + ["human_notes", "use_for_extraction"]
    )
    return pd.DataFrame(rows_out, columns=cols)


def sheet_6_extraction_issues(doc, no_image_processing: bool) -> pd.DataFrame:
    rows = []
    tables = list(getattr(doc, "tables", []) or [])

    for tbl_idx, tbl in enumerate(tables, start=1):
        pn = get_page_no(tbl)
        if not _grid_has_content(tbl):
            children = _resolve_table_group_children(doc, tbl)
            rows.append({
                "issue_type": "empty_table_grid",
                "page_number": pn,
                "table_number": tbl_idx,
                "row_number": "",
                "description": (
                    f"Table {tbl_idx} on page {pn} has an empty cell grid "
                    f"({len(children)} text fragments found in table region). "
                    "Table structure detection was "
                    + ("disabled (--no-image-processing)." if no_image_processing else "not available for this table.")
                    + " Row/column layout was not parsed. "
                    "Content is visible in sheets 2 and 3 as flat text."
                ),
                "human_notes": "",
                "resolved": "",
            })

    if no_image_processing:
        rows.append({
            "issue_type": "extraction_setting",
            "page_number": "",
            "table_number": "",
            "row_number": "",
            "description": (
                "Extraction was run with --no-image-processing. "
                "Table structure (rows/columns) was not detected by AI vision. "
                "Re-run without --no-image-processing for full table grid extraction "
                "(requires more memory and time)."
            ),
            "human_notes": "",
            "resolved": "",
        })

    return pd.DataFrame(rows, columns=[
        "issue_type", "page_number", "table_number", "row_number",
        "description", "human_notes", "resolved",
    ])


# ---------------------------------------------------------------------------
# Excel workbook formatting
# ---------------------------------------------------------------------------

REVIEW_COLUMNS = {"human_notes", "use_for_extraction", "resolved"}


def format_sheet(ws, df: pd.DataFrame, sheet_name: str):
    """Apply header formatting, column widths, and freeze pane to a worksheet."""
    # Header row formatting
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)

    # Highlight review columns
    header_names = {cell.value: cell.column for cell in ws[1]}
    review_col_indices = {
        col_idx for name, col_idx in header_names.items()
        if name in REVIEW_COLUMNS
    }
    for cell in ws[1]:
        if cell.column in review_col_indices:
            cell.fill = REVIEW_FILL
            cell.font = REVIEW_FONT

    # Auto-size columns
    col_widths = {}
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        col_name = col[0].value or ""
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=0
        )
        if col_name in REVIEW_COLUMNS:
            width = 25
        elif col_name in ("chunk_text", "text_preview", "raw_row_text", "nearby_text_preview"):
            width = min(max(max_len, 30), MAX_COL_WIDTH)
        elif col_name in ("cell_text",):
            width = min(max(max_len, 20), 60)
        else:
            width = min(max(max_len + 2, MIN_COL_WIDTH), 40)
        col_widths[col_letter] = width

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Wrap text in long-content columns
    long_cols = {"chunk_text", "text_preview", "raw_row_text", "nearby_text_preview", "cell_text"}
    long_col_indices = {
        col_idx for name, col_idx in header_names.items() if name in long_cols
    }
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in long_col_indices:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    # Freeze header row
    ws.freeze_panes = "A2"


def write_workbook(sheets: dict, out_path: Path, page_image_paths: list | None = None):
    """Write all sheets to an Excel workbook with formatting.

    If page_image_paths is provided, a '0_page_images' sheet is prepended with
    each page rendered as an embedded image.
    """
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Re-open for formatting and optional image embedding
    wb = load_workbook(out_path)

    if page_image_paths:
        ws_img = wb.create_sheet("0_page_images", 0)
        ws_img.sheet_view.showGridLines = False
        ws_img.column_dimensions["A"].width = 5   # narrow label column
        ws_img.column_dimensions["B"].width = 90  # image column

        current_row = 1
        for png_path in page_image_paths:
            if not png_path.exists():
                continue
            # Label row
            label_cell = ws_img.cell(row=current_row, column=1, value=png_path.stem.upper())
            label_cell.font = Font(bold=True, size=11)
            label_cell.alignment = Alignment(vertical="top")
            ws_img.row_dimensions[current_row].height = 18
            current_row += 1

            # Embed image — display at 680px wide, proportional height
            img = XLImage(str(png_path))
            orig_w, orig_h = img.width, img.height
            display_w = 680
            display_h = int(orig_h * display_w / orig_w) if orig_w else orig_h
            img.width = display_w
            img.height = display_h
            ws_img.add_image(img, f"B{current_row}")

            # Reserve enough rows for the image height (Excel row height in pts, 1pt ≈ 0.75px)
            row_height_pts = 0.75 * display_h
            ws_img.row_dimensions[current_row].height = row_height_pts
            current_row += 1

            # Gap row
            ws_img.row_dimensions[current_row].height = 12
            current_row += 1

    for sheet_name, df in sheets.items():
        ws = wb[sheet_name]
        format_sheet(ws, df, sheet_name)
    wb.save(out_path)


# ---------------------------------------------------------------------------
# Raw table CSV export
# ---------------------------------------------------------------------------

def write_table_csvs(doc, tables_dir: Path):
    tables = list(getattr(doc, "tables", []) or [])
    if not tables:
        return []

    tables_dir.mkdir(parents=True, exist_ok=True)
    written = []

    for tbl_idx, tbl in enumerate(tables, start=1):
        pn = get_page_no(tbl)
        csv_path = tables_dir / f"table_{tbl_idx:03d}_page{pn}.csv"

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if _grid_has_content(tbl):
                # Normal path: write grid rows
                try:
                    for row in tbl.data.grid:
                        cells = [(cell.text or "").strip() for cell in row]
                        writer.writerow(cells)
                except Exception:
                    pass
            else:
                # Fallback: write one column of text fragments from the group
                writer.writerow(["# no_grid_structure — table detection was off"])
                for item in _resolve_table_group_children(doc, tbl):
                    writer.writerow([safe_text(item)])

        written.append(csv_path.name)

    return written


# ---------------------------------------------------------------------------
# Run summary
# ---------------------------------------------------------------------------

def write_run_summary(
    out_dir: Path,
    input_path: Path,
    selected_pages: list | None,
    page_count,
    extracted_pages_count,
    extracted_pages_max,
    coverage_ok: bool,
    table_count: int,
    chunk_count: int,
    heading_count: int,
    table_csv_names: list,
    image_names: list,
    ts: str,
    no_image_processing: bool,
    allow_partial: bool,
):
    lines = [
        "run_summary.txt",
        "=" * 60,
        f"Input PDF       : {input_path}",
        f"Run timestamp   : {ts}",
        f"Output folder   : {out_dir}",
        f"Selected pages  : {selected_pages if selected_pages else 'all'}",
        "",
        "Docling settings:",
        f"  no_image_processing : {no_image_processing}",
        f"  allow_partial       : {allow_partial}",
        "",
        "Extraction results:",
        f"  PDF pages declared  : {page_count}",
        f"  Pages with content  : {extracted_pages_count}",
        f"  Highest page found  : {extracted_pages_max}",
        f"  Coverage OK         : {coverage_ok}",
        f"  Tables found        : {table_count}",
        f"  Text chunks         : {chunk_count}",
        f"  Headings            : {heading_count}",
        "",
        "Output files:",
        "  review_workbook.xlsx",
        "  docling_document.json",
        "  output.md",
    ]
    if image_names:
        lines.append(f"  images/ ({len(image_names)} PNG files):")
        for name in image_names:
            lines.append(f"    {name}")
    if table_csv_names:
        lines.append(f"  tables/ ({len(table_csv_names)} CSV files):")
        for name in table_csv_names:
            lines.append(f"    {name}")
    lines += [
        "  run_summary.txt",
        "",
        "Workbook sheets:",
        "  0_page_images      — rendered page images (if --pages used)",
        "  1_source_pages     — one row per PDF page",
        "  2_raw_tables       — one row per table cell (Docling grid)",
        "  3_raw_table_rows   — one row per table row, cells spread across columns",
        "  3b_bbox_rows       — bbox-reconstructed rows (useful when table detection missed)",
        "  4_raw_chunks       — non-table text chunks (paragraphs, captions, footnotes ...)",
        "  5_headings_index   — all detected headings with nearby text preview",
        "  6_extraction_issues — blank; fill in during review",
        "",
        "Human review columns (highlighted yellow):",
        "  use_for_extraction  : leave blank, or enter  yes / no",
        "  resolved            : leave blank, or enter  yes / no  (issues sheet only)",
        "  human_notes         : free text notes",
    ]
    summary_path = out_dir / "run_summary.txt"
    summary_path.write_text("\n".join(lines), encoding="utf-8")
    return summary_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Extract a PDF with Docling and produce a human-reviewable Excel workbook. "
            "No LLM calls. No schema mapping. Local only."
        )
    )
    parser.add_argument("--input", required=True, help="Path to the PDF file")
    parser.add_argument(
        "--output-folder",
        default=None,
        metavar="DIR",
        help=(
            "Override the full output directory path. "
            "Must be under .local/extraction-runs/."
        ),
    )
    parser.add_argument(
        "--run-name",
        default=None,
        metavar="NAME",
        help=(
            "Name for the run subfolder under .local/extraction-runs/. "
            "Defaults to <pdf-stem>_<timestamp>."
        ),
    )
    parser.add_argument(
        "--no-image-processing",
        action="store_true",
        default=False,
        help=(
            "Skip image-based layout analysis and table structure detection. "
            "Prevents memory crashes on large PDFs. Recommended for first runs."
        ),
    )
    parser.add_argument(
        "--pages",
        default=None,
        metavar="PAGES",
        help=(
            "Extract only specific pages, e.g. '5-7' or '5,6,7'. "
            "Creates a sub-PDF for Docling and renders page images. "
            "Page numbers are 1-indexed (matching PDF page numbers)."
        ),
    )
    parser.add_argument(
        "--allow-partial",
        action="store_true",
        default=False,
        help="Allow extraction even if Docling did not reach 90%% of PDF pages.",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        sys.exit(f"[ERROR] Input file not found: {input_path}")
    if input_path.suffix.lower() != ".pdf":
        sys.exit(f"[ERROR] Input must be a .pdf file. Got: {input_path.name}")

    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")

    if args.output_folder:
        out_dir = Path(args.output_folder)
        resolved_root = SAFE_OUTPUT_ROOT.resolve()
        resolved_out = out_dir.resolve()
        try:
            resolved_out.relative_to(resolved_root)
        except ValueError:
            sys.exit(
                f"[ERROR] --output-folder must be under {SAFE_OUTPUT_ROOT}/.\nGot: {out_dir}"
            )
    else:
        out_dir = resolve_output_dir(args.run_name, input_path.stem, ts)

    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"[workbook] Output folder: {out_dir}")

    # --- Page selection ---
    selected_pages = None
    docling_input = input_path
    page_offset = 0  # added to all Docling page numbers to recover original page numbers

    if args.pages:
        selected_pages = parse_pages_arg(args.pages)
        print(f"[workbook] Selected pages: {selected_pages}")

        # Extract sub-PDF for Docling
        sub_pdf_path = out_dir / f"sub_pages_{'_'.join(str(p) for p in selected_pages)}.pdf"
        try:
            docling_input = extract_pages_to_pdf(input_path, selected_pages, sub_pdf_path)
        except Exception as exc:
            sys.exit(f"[ERROR] Could not extract sub-PDF: {exc}")

        # Page offset: Docling will see pages 1..N internally; add offset to restore originals
        page_offset = selected_pages[0] - 1

        # Render page images from the original PDF
        images_dir = out_dir / "images"
        try:
            page_image_paths = render_page_images(input_path, selected_pages, images_dir, scale=2.0)
        except Exception as exc:
            print(f"[workbook] WARNING: Image rendering failed: {exc}")
            page_image_paths = []
    else:
        page_image_paths = []

    # --- Docling extraction ---
    try:
        doc, page_count, extracted_pages_count, extracted_pages_max, coverage_ok = run_docling(
            docling_input,
            no_image_processing=args.no_image_processing,
            allow_partial=args.allow_partial,
            out_dir=out_dir,
        )
    except ImportError as exc:
        sys.exit(
            f"[ERROR] Docling import failed: {exc}\n\n"
            "Activate the venv first:\n"
            "  .venv-docling\\Scripts\\Activate.ps1\n"
        )
    except Exception as exc:
        sys.exit(f"[ERROR] Docling extraction failed: {exc}")

    print("[workbook] Building data frames ...")

    # --- Raw Docling outputs ---
    try:
        md_text = doc.export_to_markdown()
        (out_dir / "output.md").write_text(md_text, encoding="utf-8")
        print("[workbook] Wrote output.md")
    except Exception as exc:
        print(f"[workbook] WARNING: Markdown export failed: {exc}")

    try:
        (out_dir / "docling_document.json").write_text(doc_to_json_str(doc), encoding="utf-8")
        print("[workbook] Wrote docling_document.json")
    except Exception as exc:
        print(f"[workbook] WARNING: JSON export failed: {exc}")

    # --- Build data frames ---
    page_items = build_page_items(doc)
    page_heading_map = build_page_heading_map(page_items)
    table_child_refs = build_table_child_refs(doc)

    df1 = sheet_1_source_pages(doc, page_items, page_heading_map)
    df2 = sheet_2_raw_tables(doc)
    df3 = sheet_3_raw_table_rows(doc)
    df3b = sheet_3b_bbox_rows(doc, page_offset=0)  # offset applied below after build
    df4 = sheet_4_raw_chunks(doc, page_items, page_heading_map, table_child_refs)
    df5 = sheet_5_headings_index(doc, page_items)
    df6 = sheet_6_extraction_issues(doc, no_image_processing=args.no_image_processing)

    # Remap page numbers to original PDF page numbers when a sub-PDF was used
    if page_offset > 0:
        def _add_offset(x):
            try:
                return int(x) + page_offset if x != "" and pd.notna(x) else x
            except (ValueError, TypeError):
                return x
        for df in (df1, df2, df3, df3b, df4, df5, df6):
            if "page_number" in df.columns:
                df["page_number"] = df["page_number"].apply(_add_offset)

    table_count = len(list(getattr(doc, "tables", []) or []))
    chunk_count = len(df4)
    heading_count = len(df5)

    sheets = {
        "1_source_pages": df1,
        "2_raw_tables": df2,
        "3_raw_table_rows": df3,
        "3b_bbox_rows": df3b,
        "4_raw_chunks": df4,
        "5_headings_index": df5,
        "6_extraction_issues": df6,
    }

    # --- Table CSVs ---
    tables_dir = out_dir / "tables"
    table_csv_names = write_table_csvs(doc, tables_dir)
    if table_csv_names:
        print(f"[workbook] Wrote {len(table_csv_names)} table CSV(s) to tables/")

    # --- Workbook (with optional image sheet) ---
    workbook_path = out_dir / "review_workbook.xlsx"
    print("[workbook] Writing review_workbook.xlsx ...")
    write_workbook(sheets, workbook_path, page_image_paths=page_image_paths or None)
    print(f"[workbook] Wrote review_workbook.xlsx ({workbook_path.stat().st_size:,} bytes)")

    # --- Run summary ---
    image_names = [p.name for p in page_image_paths]
    write_run_summary(
        out_dir=out_dir,
        input_path=input_path,
        selected_pages=selected_pages,
        page_count=page_count,
        extracted_pages_count=extracted_pages_count,
        extracted_pages_max=extracted_pages_max,
        coverage_ok=coverage_ok,
        table_count=table_count,
        chunk_count=chunk_count,
        heading_count=heading_count,
        table_csv_names=table_csv_names,
        image_names=image_names,
        ts=ts,
        no_image_processing=args.no_image_processing,
        allow_partial=args.allow_partial,
    )
    print("[workbook] Wrote run_summary.txt")

    # --- Final report ---
    print()
    print("=" * 60)
    print("Review workbook complete.")
    print(f"  Output folder : {out_dir}")
    if selected_pages:
        print(f"  Selected pages: {selected_pages}")
    print(f"  Pages         : {page_count}  (extracted content on {extracted_pages_count} pages)")
    print(f"  Tables        : {table_count}")
    print(f"  Text chunks   : {chunk_count}")
    print(f"  Headings      : {heading_count}")
    print()
    print("Files written:")
    print(f"  {workbook_path.name}")
    print(f"  docling_document.json")
    print(f"  output.md")
    for name in image_names:
        print(f"  images/{name}")
    for name in table_csv_names:
        print(f"  tables/{name}")
    print(f"  run_summary.txt")
    print("=" * 60)

    if not coverage_ok:
        print("\n[WARNING] Extraction coverage was below 90%. Review run_summary.txt.")
        sys.exit(1)


if __name__ == "__main__":
    main()
